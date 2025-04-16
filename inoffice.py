#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys, os
import json
import re
import argparse
import datetime
import pandas as pd
import jpholiday
import locale
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.styles import Border, Side

# 存储数据的 JSON 文件
DATA_FILE = f"/home/r/proj/zz/script/inoffice-{datetime.datetime.now().year}.json"
BREAK_START = "12:30"
BREAK_END = "13:30"
ENV_NAME = 'OUTPUT_PREFIX'
OUTPUT = "{oprefix}出勤_{target_month:02d}.xlsx"
HOURS = 8.0

# 读取 JSON 数据
def load_data():
    global DATA_FILE
    if not os.path.exists(os.path.dirname(DATA_FILE)):
        print(f"[{os.path.dirname(DATA_FILE)}]不存在，使用用户主目录")
        DATA_FILE = os.path.join(os.path.expanduser("~"), f"inoffice-{datetime.datetime.now().year}.json")
    #print(f"读取文件[{DATA_FILE}]")
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {}

# 保存 JSON 数据
def save_data(data):
    global DATA_FILE
    if not os.path.exists(os.path.dirname(DATA_FILE)):
        DATA_FILE = os.path.join(os.path.expanduser("~"), f"inoffice-{datetime.datetime.now().year}.json")
    #print(f"保存文件[{DATA_FILE}]")
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=4)

def get_weekday(dateobj):
    locale.setlocale(locale.LC_TIME, 'ja_JP.UTF-8')
    return dateobj.strftime('%a')

# 记录考勤
def record_attendance(month, day, times):
    data = load_data()
    today = f"{month:02d}-{day:02d}"
    if today not in data:
        data[today] = {"times": [], "memo": ""}
    for time in times:
        data[today]["times"].append(time)
    data[today]["times"] = sorted(set(data[today]["times"]))  # 去重并排序
    save_data(data)
    print(f"已记录：{today} {times}")

# 记录备注
def record_memo(month, day, memo):
    data = load_data()
    today = f"{month:02d}-{day:02d}"
    if today not in data:
        data[today] = {"times": [], "memo": ""}
    data[today]["memo"] = memo
    save_data(data)
    print(f"已更新备注：{today} -> {memo}")

# 计算出勤时间
def round_time_to_nearest_10(time_str, round_up=False):
    """将时间对齐到 10 分钟单位"""
    h, m = map(int, time_str.split(":"))
    if round_up:
        m = ((m + 9) // 10) * 10
    else:
        m = (m // 10) * 10
    if m >= 60:
        h += 1
        m = 0
    return f"{h:02d}:{m:02d}"

# 生成 Excel
def export_to_excel(target_month):
    data = load_data()
    today = datetime.date.today()
    year = today.year

    days_in_month = (datetime.date(year + (target_month // 12), target_month % 12 + 1, 1) - datetime.timedelta(days=1)).day
    records = []
    total_hours = 0

    for day in range(1, days_in_month + 1):
        date_str = f"{target_month:02d}-{day:02d}"
        date_obj = datetime.date(year, target_month, day)
        info = data.get(date_str, {"times": [], "memo": ""})

        if info["times"]:
            if len(info["times"]) < 2:
                print('记录不全:')
                print(' '.join([date_str, info['memo'], 'times:', str(info['times'])]))
                sys.exit(1)
            l = [f'0{time}' if len(time) == 4 else time for time in info["times"]]
            assert all(map(lambda s: len(s) == 5, l))
            l.sort()
            start_time = round_time_to_nearest_10(l[0], round_up=False)
            end_time = round_time_to_nearest_10(l[-1], round_up=True)
        else:
            start_time, end_time = "", ""

        work_hours = calculate_work_hours(start_time, end_time)
        if work_hours:
            total_hours += work_hours

        records.append([f"{day:02d}日({get_weekday(date_obj)})", start_time, end_time, work_hours, info["memo"]])

    df = pd.DataFrame(records, columns=["日期", "上班时间", "下班时间", "出勤时长", "备注"])

    # **周末 & 日本法定假日 & 工时不满高亮**
    fill_holiday = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
    fill_weekend = PatternFill(start_color="DDDDFF", end_color="DDDDFF", fill_type="solid")
    fill_bad = PatternFill(start_color="880000", end_color="880000", fill_type="solid")

    oprefix = os.getenv(ENV_NAME) or ''
    file_name = eval(f'f"{OUTPUT}"')
    if not os.path.dirname(file_name):
        # 如果没有路径，则在前面加上 './'
        file_name = f"./{file_name}"
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=f"{target_month}月", startrow=4)

        sheet = writer.sheets[f"{target_month}月"]

        # **合并月份标题**
        sheet.merge_cells("A1:E1")
        sheet["A1"] = f"{target_month} 月 出勤记录"
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # **合并总时长标题**
        sheet.merge_cells("A2:E2")
        sheet["A2"] = f"总出勤时长: {total_hours:.1f} 小时"
        sheet["A2"].font = Font(size=14, bold=True)
        sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # **重新填充表头**
        sheet["A5"] = "日期"

        # **添加计算说明**
        sheet[f"A{len(df)+7}"] = f"※ {BREAK_START} - {BREAK_END} 的休息时间不计入出勤时长"

        # **统一格式**
        content_font = Font(size=11)
        for col in "ABCDE":
            for row in range(1, len(df) + 7):
                cell = sheet[f"{col}{row}"]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = content_font

        # **设置列宽**
        column_widths = {"A": 10, "B": 12, "C": 12, "D": 12, "E": 20}
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width

        for idx, row in enumerate(df.itertuples(), start=6):
            #date_obj = datetime.date(year, target_month, int(row.日期[:-1]))
            target_day = int(re.search(r'\d+', row.日期).group())
            target_hours = row.出勤时长
            date_obj = datetime.date(year, target_month, target_day)
            if jpholiday.is_holiday(date_obj):
                sheet[f"E{idx}"] = jpholiday.is_holiday_name(date_obj)
                for col in "ABCDE":
                    sheet[f"{col}{idx}"].fill = fill_holiday
            elif date_obj.weekday() in [5, 6]:  # 周六、周日
                for col in "ABCDE":
                    sheet[f"{col}{idx}"].fill = fill_weekend
            #elif target_hours < HOURS:
            #    for col in "ABCDE":
            #        sheet[f"{col}{idx}"].fill = fill_bad

        # **边框样式**
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for row in sheet.iter_rows(min_row=5, max_row=len(df) + 5, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border

    print(f"环境变量OUTPUT_PREFIX：{oprefix}")
    print(f"Excel 导出完成：{file_name}")

# 计算出勤时长
def calculate_work_hours(start, end):
    if not start or not end:
        return ""

    start_time = datetime.datetime.strptime(start, "%H:%M")
    end_time = datetime.datetime.strptime(end, "%H:%M")

    # 休息时间段
    break_start = datetime.datetime.strptime(BREAK_START, "%H:%M")
    break_end = datetime.datetime.strptime(BREAK_END, "%H:%M")

    work_seconds = (end_time - start_time).total_seconds()

    if start_time < break_end and end_time > break_start:
        work_seconds -= max(0, (min(end_time, break_end) - max(start_time, break_start)).total_seconds())

    work_hours = work_seconds / 3600
    return max(0, round(work_hours, 1))

def fill_workdays(start_time, end_time, inmonth):
    data = load_data()
    today = datetime.date.today()
    year, month = today.year, inmonth or today.month
    days_in_month = (datetime.date(year, month % 12 + 1, 1) - datetime.timedelta(days=1)).day

    for day in range(1, days_in_month + 1):
        date_str = f"{month:02d}-{day:02d}"
        date_obj = datetime.date(year, month, day)

        if date_str not in data and date_obj.weekday() < 5 and not jpholiday.is_holiday(date_obj):
            data[date_str] = {"times": [start_time, end_time], "memo": ""}

    save_data(data)
    print(f"已填充 {month} 月所有未记录的工作日：{start_time} - {end_time}")

#        date_str = f"{target_month:02d}-{day:02d}"
def list_workdays(target_month):
    data = load_data()

    print(f"{target_month} 月记录：")
    for k, v in data.items():
        prefix = f"{target_month:02d}-"
        if k.startswith(prefix):
            print(' '.join([k, v['memo'], 'times:', str(v['times'])]))


if __name__ == "__main__":
    # 解析命令行参数
    parser = argparse.ArgumentParser(description="出勤记录工具。不带参数运行会进行打卡。")
    parser.add_argument("-m", nargs="+", metavar="MEMO", help="记录备注")
    parser.add_argument("-s", nargs="+", metavar=("DATE"), help="记录指定日期的时间")
    parser.add_argument("-o", action="store_true", help="导出当月 Excel")
    parser.add_argument("-M", type=int, metavar="MONTH", help="导出指定月份的 Excel")
    parser.add_argument("--fill", nargs=2, metavar=("START", "END"), help="填充所有未记录的工作日时间")
    parser.add_argument("-l", "--list", dest="list", action="store_true", help="列出所有记录")

    args = parser.parse_args()

    # 获取当前日期
    now = datetime.datetime.now()

    #print('--------args.m:', args.m)
    #print('--------args.s:', args.s)
    #print('--------args.o:', args.o)
    #print('--------args.M:', args.M)
    # 处理参数逻辑
    if args.m:
        memo_text = " ".join(args.m)
        if args.s:
            record_memo(int(args.s[0]), int(args.s[1]), memo_text)
        else:
            record_memo(now.month, now.day, memo_text)
    elif args.o:
        export_to_excel(args.M if args.M else now.month)
    elif args.fill:
        fill_workdays(args.fill[0], args.fill[1], int(args.s[0]) if args.s else None)
    elif args.list:
        list_workdays(args.M if args.M else now.month)
    else:
        if args.s:
            record_attendance(int(args.s[0]), int(args.s[1]), args.s[2:])
        else:
            record_attendance(now.month, now.day, [now.strftime("%H:%M")])
